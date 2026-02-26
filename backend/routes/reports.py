# backend/routes/reports.py
"""
Маршруты для отчётов и экспорта
"""
from fastapi import APIRouter, Depends, HTTPException, status, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import datetime, timedelta
import io
import csv

from database import get_db
from models import BusStop, User, AuditLog
from schemas import StatsResponse, ReportFilter
from core.dependencies import get_current_user, require_any_role
from middleware.audit import AuditLogger


router = APIRouter(tags=["Отчёты"])


def get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


@router.get("/dashboard")
async def get_dashboard_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_any_role)
):
    """
    Данные для дашборда
    """
    # Общая статистика
    total = db.query(BusStop).count()
    active = db.query(BusStop).filter(BusStop.status == "active").count()
    repair = db.query(BusStop).filter(BusStop.status == "repair").count()
    dismantled = db.query(BusStop).filter(BusStop.status == "dismantled").count()
    inactive = db.query(BusStop).filter(BusStop.status == "inactive").count()
    
    # По состоянию
    excellent = db.query(BusStop).filter(BusStop.condition == "excellent").count()
    satisfactory = db.query(BusStop).filter(BusStop.condition == "satisfactory").count()
    needs_repair = db.query(BusStop).filter(BusStop.condition == "needs_repair").count()
    critical = db.query(BusStop).filter(BusStop.condition == "critical").count()
    
    # Проверено за этот месяц
    first_day = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    inspected = db.query(BusStop).filter(
        BusStop.last_inspection_date >= first_day
    ).count()
    
    # По районам
    districts = db.query(
        BusStop.district,
        func.count(BusStop.id)
    ).group_by(BusStop.district).all()
    
    by_district = [{"name": d[0], "value": d[1]} for d in districts]
    
    # Требуют внимания (критическое или требует ремонта)
    attention_needed = db.query(BusStop).filter(
        BusStop.condition.in_(["critical", "needs_repair"])
    ).order_by(
        BusStop.condition.desc(),
        BusStop.updated_at.desc()
    ).limit(10).all()
    
    return {
        "stats": {
            "total": total,
            "active": active,
            "repair": repair,
            "dismantled": dismantled,
            "inactive": inactive,
            "excellent": excellent,
            "satisfactory": satisfactory,
            "needs_repair": needs_repair,
            "critical": critical,
            "inspected_this_month": inspected
        },
        "by_district": by_district,
        "by_status": [
            {"name": "Активна", "value": active, "color": "#22c55e"},
            {"name": "В ремонте", "value": repair, "color": "#eab308"},
            {"name": "Демонтирована", "value": dismantled, "color": "#6b7280"},
            {"name": "Недоступна", "value": inactive, "color": "#ef4444"}
        ],
        "by_condition": [
            {"name": "Отличное", "value": excellent, "color": "#22c55e"},
            {"name": "Удовлетворительное", "value": satisfactory, "color": "#3b82f6"},
            {"name": "Требует ремонта", "value": needs_repair, "color": "#f97316"},
            {"name": "Критическое", "value": critical, "color": "#ef4444"}
        ],
        "attention_needed": [
            {
                "id": s.id,
                "stop_id": s.stop_id,
                "address": s.address,
                "district": s.district,
                "condition": s.condition.value if s.condition else None,
                "status": s.status.value if s.status else None
            }
            for s in attention_needed
        ]
    }


@router.get("/export")
async def export_report(
    request: Request,
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    district: Optional[str] = None,
    status: Optional[str] = None,
    condition: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_any_role)
):
    """
    Экспорт данных в Excel/CSV
    """
    # Формируем запрос
    query = db.query(BusStop)
    
    if district:
        query = query.filter(BusStop.district == district)
    
    if status:
        query = query.filter(BusStop.status == status)
    
    if condition:
        query = query.filter(BusStop.condition == condition)
    
    stops = query.order_by(BusStop.stop_id).all()
    
    # Логируем экспорт
    AuditLogger.log_export(
        db=db,
        user=current_user,
        export_type=format,
        filters={
            "district": district,
            "status": status,
            "condition": condition,
            "count": len(stops)
        },
        ip_address=get_client_ip(request)
    )
    
    if format == "csv":
        return export_csv(stops)
    else:
        return export_xlsx(stops)


def export_csv(stops):
    """Экспорт в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    headers = [
        "ID", "№ Паспорта", "Адрес", "Ориентир", "Район", "Маршруты",
        "Широта", "Долгота", "Статус", "Состояние", "Соответствие нормам",
        "Тип", "Кол-во стоек", "Год постройки", "Цвет",
        "Состояние сидений", "Тип крыши", "Состояние крыши",
        "Электропитание", "Урна", "Последняя проверка", "Инспектор"
    ]
    writer.writerow(headers)
    
    # Данные
    status_labels = {
        "active": "Активна",
        "repair": "В ремонте",
        "dismantled": "Демонтирована",
        "inactive": "Недоступна"
    }
    
    condition_labels = {
        "excellent": "Отличное",
        "satisfactory": "Удовлетворительное",
        "needs_repair": "Требует ремонта",
        "critical": "Критическое"
    }
    
    for stop in stops:
        row = [
            stop.stop_id,
            stop.passport_number or "",
            stop.address,
            stop.landmark or "",
            stop.district,
            stop.routes or "",
            stop.latitude,
            stop.longitude,
            status_labels.get(stop.status.value, stop.status.value) if stop.status else "",
            condition_labels.get(stop.condition.value, stop.condition.value) if stop.condition else "",
            "Да" if stop.meets_standards else "Нет",
            stop.stop_type.value if stop.stop_type else "",
            stop.legs_count,
            stop.year_built or "",
            stop.paint_color or "",
            condition_labels.get(stop.seats_condition.value, "") if stop.seats_condition else "",
            stop.roof_type.value if stop.roof_type else "",
            condition_labels.get(stop.roof_condition.value, "") if stop.roof_condition else "",
            "Да" if stop.has_electricity else "Нет",
            "Да" if stop.has_bin else "Нет",
            stop.last_inspection_date.strftime("%d.%m.%Y") if stop.last_inspection_date else "",
            stop.inspector_name or ""
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=bus_stops_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


def export_xlsx(stops):
    """Экспорт в Excel (XLSX)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Модуль openpyxl не установлен"
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Остановки"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        "ID", "№ Паспорта", "Адрес", "Ориентир", "Район", "Маршруты",
        "Широта", "Долгота", "Статус", "Состояние", "Соответствие нормам",
        "Тип", "Кол-во стоек", "Год постройки", "Цвет",
        "Состояние сидений", "Тип крыши", "Состояние крыши",
        "Электропитание", "Урна", "Последняя проверка", "Инспектор"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Данные
    status_labels = {
        "active": "Активна",
        "repair": "В ремонте",
        "dismantled": "Демонтирована",
        "inactive": "Недоступна"
    }
    
    condition_labels = {
        "excellent": "Отличное",
        "satisfactory": "Удовлетворительное",
        "needs_repair": "Требует ремонта",
        "critical": "Критическое"
    }
    
    for row_num, stop in enumerate(stops, 2):
        data = [
            stop.stop_id,
            stop.passport_number or "",
            stop.address,
            stop.landmark or "",
            stop.district,
            stop.routes or "",
            stop.latitude,
            stop.longitude,
            status_labels.get(stop.status.value, stop.status.value) if stop.status else "",
            condition_labels.get(stop.condition.value, stop.condition.value) if stop.condition else "",
            "Да" if stop.meets_standards else "Нет",
            stop.stop_type.value if stop.stop_type else "",
            stop.legs_count,
            stop.year_built or "",
            stop.paint_color or "",
            condition_labels.get(stop.seats_condition.value, "") if stop.seats_condition else "",
            stop.roof_type.value if stop.roof_type else "",
            condition_labels.get(stop.roof_condition.value, "") if stop.roof_condition else "",
            "Да" if stop.has_electricity else "Нет",
            "Да" if stop.has_bin else "Нет",
            stop.last_inspection_date.strftime("%d.%m.%Y") if stop.last_inspection_date else "",
            stop.inspector_name or ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Закрепляем заголовок
    ws.freeze_panes = "A2"
    
    # Сохраняем в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=bus_stops_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )


@router.get("/audit-log")
async def get_audit_log(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_any_role)
):
    """
    Журнал аудита (только для админа)
    """
    if current_user.role.value != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Доступ запрещён"
        )
    
    query = db.query(AuditLog)
    
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    
    if action:
        query = query.filter(AuditLog.action == action)
    
    if resource_type:
        query = query.filter(AuditLog.resource_type == resource_type)
    
    if date_from:
        query = query.filter(AuditLog.timestamp >= date_from)
    
    if date_to:
        query = query.filter(AuditLog.timestamp <= date_to)
    
    total = query.count()
    
    logs = query.order_by(AuditLog.timestamp.desc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()
    
    return {
        "logs": [
            {
                "id": log.id,
                "user_id": log.user_id,
                "user_email": log.user_email,
                "action": log.action,
                "resource_type": log.resource_type,
                "resource_id": log.resource_id,
                "details": log.details,
                "ip_address": log.ip_address,
                "timestamp": log.timestamp
            }
            for log in logs
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }
